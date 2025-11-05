import time
import pandas as pd
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoAlertPresentException, UnexpectedAlertPresentException

# Configurations
excel_file_path = 'test_data.xlsx'  # Excel file
sheet_name = 'Лист1'  # Sheet name
url = 'https://qa-ep-bva-practice-assignment.vercel.app/'

# Load input data from Excel column B (index 1)
df = pd.read_excel(excel_file_path, sheet_name=sheet_name)
input_values = df.iloc[:, 1].tolist()

# Setup Selenium WebDriver (make sure chromedriver is installed and in PATH)
chrome_options = Options()
chrome_options.add_argument('--headless')  # Run without browser UI
service = Service()
driver = webdriver.Chrome(service=service, options=chrome_options)
wait = WebDriverWait(driver, 10)

results = []

try:
    for test_value in input_values:
        driver.get(url)

        # Locate the Implementation 8 form container
        implementation_div = wait.until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "div[data-testid='Implementation 11']"))
        )

        # Input the test value
        input_box = implementation_div.find_element(By.CSS_SELECTOR, "input.form-control")
        input_box.clear()
        input_box.send_keys(str(test_value))

        # Click Validate button
        validate_button = implementation_div.find_element(By.CSS_SELECTOR, "button.btn-primary")
        validate_button.click()

        # Handle possible alert first with short wait
        try:
            wait_short = WebDriverWait(driver, 2)
            alert = wait_short.until(EC.alert_is_present())
            alert_text = alert.text
            alert.accept()
            message_text = alert_text
        except TimeoutException:
            # No alert, try to get toast message
            try:
                toast_container = wait.until(
                    EC.visibility_of_element_located((By.CSS_SELECTOR, "div.Toastify"))
                )
                toast_message_element = wait.until(
                    EC.visibility_of_element_located((By.CSS_SELECTOR, "div.Toastify__toast-body > div:last-child"))
                )
                message_text = toast_message_element.text.strip()
            except TimeoutException:
                message_text = "No toast or alert message found"

        print(f"Test value: {test_value} => Result: {message_text}")
        results.append(message_text)

        time.sleep(1)

except UnexpectedAlertPresentException as e:
    # Catch any unexpected alerts during operation
    alert = driver.switch_to.alert
    alert_text = alert.text
    alert.accept()
    print(f"Handled unexpected alert: {alert_text}")
    results.append(alert_text)

except Exception as e:
    print("Error during automation:", e)

finally:
    driver.quit()

# Write results back to Excel column C (index 2)
wb = load_workbook(excel_file_path)
ws = wb[sheet_name]

for idx, result in enumerate(results, start=2):  # Assuming header is row 1
    ws.cell(row=idx, column=3, value=result)

wb.save(excel_file_path)
print("Excel file updated with results successfully.")
